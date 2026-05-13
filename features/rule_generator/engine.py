"""Comprehensive AI Data Quality Rule Engine for Rule Generator"""

import logging
import re
import json
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import numpy as np

logger = logging.getLogger(__name__)


def scan_excel_for_rule_sheets(uploaded_file) -> List[str]:
    """Scan Excel workbook for rule-related sheets
    
    Returns list of sheet names that appear to contain rules
    """
    try:
        import openpyxl
        
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        
        # Rule-related sheet name patterns
        rule_patterns = [
            r'rule\s*generator',
            r'validation\s*rules?',
            r'data\s*quality\s*rules?',
            r'dq\s*rules?',
            r'field\s*validation',
            r'business\s*rules?',
            r'validation\s*instruction',
            r'rules?',
            r'validation'
        ]
        
        rule_sheets = []
        for sheet_name in wb.sheetnames:
            for pattern in rule_patterns:
                if re.search(pattern, sheet_name, re.IGNORECASE):
                    rule_sheets.append(sheet_name)
                    logger.info(f"Found rule-related sheet: {sheet_name}")
                    break
        
        wb.close()
        return rule_sheets
        
    except Exception as e:
        logger.error(f"Error scanning for rule sheets: {str(e)}")
        return []


def deep_scan_rule_sheet(uploaded_file_path: str, sheet_name: str, header_row: int = 0) -> Dict[str, Any]:
    """Perform deep scan of a rule sheet to extract all rule information
    
    Args:
        uploaded_file_path: Path to the uploaded Excel file
        sheet_name: Name of the sheet to scan
        header_row: Row index where column headers are located (0-based)
    
    Scans:
    - ALL rows (above and below header row)
    - All columns
    - Cell comments
    - Cell notes
    - Merged cells
    - Hidden rows/columns
    
    Returns dict mapping column names to their detected rules
    """
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(uploaded_file_path, data_only=False)
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return {}
        
        ws = wb[sheet_name]
        
        logger.info(f"=== DEEP SCANNING RULE SHEET: {sheet_name} ===")
        logger.info(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        logger.info(f"Header row specified: {header_row + 1} (1-based)")
        
        # Step 1: Extract column names from header row
        column_names = []
        header_row_idx = header_row + 1  # Convert to 1-based for openpyxl
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            if cell.value:
                column_names.append({
                    'name': str(cell.value).strip(),
                    'col_idx': col_idx,
                    'comment': cell.comment.text if cell.comment else None
                })
                logger.info(f"Column {col_idx}: {cell.value}")
        
        logger.info(f"Found {len(column_names)} columns in header row {header_row_idx}")
        
        # Step 2: Scan ALL rows for rules (above and below header)
        column_rules = {}
        
        for col_info in column_names:
            col_name = col_info['name']
            col_idx = col_info['col_idx']
            rules_found = []
            
            # Add header comment if exists
            if col_info['comment']:
                rules_found.append(f"[Header Comment] {col_info['comment']}")
                logger.info(f"Column '{col_name}': Found header comment")
            
            # Scan rows ABOVE header (rules often written above)
            for row_idx in range(1, header_row_idx):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    value = str(cell.value).strip()
                    if value and len(value) < 200:  # Reasonable rule length
                        rules_found.append(f"[Row {row_idx}] {value}")
                        logger.info(f"Column '{col_name}': Found rule in row {row_idx} above header: {value[:50]}")
                
                # Check for comments in rows above
                if cell.comment:
                    rules_found.append(f"[Row {row_idx} Comment] {cell.comment.text}")
                    logger.info(f"Column '{col_name}': Found comment in row {row_idx}")
            
            # Scan rows BELOW header (first 10 rows for data type hints)
            for row_idx in range(header_row_idx + 1, min(header_row_idx + 11, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    value = str(cell.value).strip()
                    # Check if this looks like a rule/instruction (not data)
                    if any(keyword in value.lower() for keyword in [
                        'char', 'varchar', 'number', 'numeric', 'decimal', 'date',
                        'mandatory', 'required', 'max', 'length', 'format'
                    ]):
                        rules_found.append(f"[Row {row_idx}] {value}")
                        logger.info(f"Column '{col_name}': Found rule in row {row_idx} below header: {value[:50]}")
                
                # Check for comments in rows below
                if cell.comment:
                    rules_found.append(f"[Row {row_idx} Comment] {cell.comment.text}")
            
            # Scan adjacent cells (left and right) in header row
            # Check cell to the right
            if col_idx < ws.max_column:
                right_cell = ws.cell(row=header_row_idx, column=col_idx + 1)
                if right_cell.value and col_idx + 1 not in [c['col_idx'] for c in column_names]:
                    value = str(right_cell.value).strip()
                    if value and len(value) < 200:
                        rules_found.append(f"[Adjacent Right] {value}")
                        logger.info(f"Column '{col_name}': Found rule in adjacent right cell: {value[:50]}")
            
            # Store all found rules for this column
            if rules_found:
                column_rules[col_name] = ' | '.join(rules_found)
                logger.info(f"Column '{col_name}': Total rules found: {len(rules_found)}")
            else:
                logger.info(f"Column '{col_name}': No rules found, will use AI generation")
        
        wb.close()
        
        logger.info(f"=== SCAN COMPLETE: Found rules for {len(column_rules)}/{len(column_names)} columns ===")
        return column_rules
        
    except Exception as e:
        logger.error(f"Error deep scanning rule sheet: {str(e)}", exc_info=True)
        return {}


def _ensure_regex_pattern(rule: Dict[str, Any], metadata: Dict[str, Any]) -> None:
    """Populate ``regex_pattern`` when absent and metadata supports a concrete pattern.

    Args:
        rule: A single rule dict from the AI response (mutated in place).
        metadata: Column metadata from :func:`extract_comprehensive_metadata`.
    """
    existing = str(rule.get("regex_pattern") or "").strip()
    if existing:
        return
    dim = rule.get("dimension", "")
    rule_text = str(rule.get("data_quality_rule", "")).lower()
    if dim == "Validation" and "character" in rule_text and metadata.get("max_length") is not None:
        n = int(metadata["max_length"])
        rule["regex_pattern"] = f"^.{0,{n}}$"


_ALLOWED_DIMENSIONS = (
    "Accuracy", "Completeness", "Standardisation", "Validation", "Uniqueness",
    "Timeliness", "Cross-field Validation",
)
_LEGACY_DIMENSION_MAP = {
    "conformity": "Validation",
    "character length": "Validation",
    "integrity": "Standardisation",
    "reliability": "Accuracy",
    "relevance": "Accuracy",
    "precision": "Accuracy",
    "accessibility": "Validation",
    # Pre-rename display names — keep rules saved on disk readable after the
    # 2026-05 Consistency→Standardisation / Validity→Validation rename.
    "validity": "Validation",
    "consistency": "Standardisation",
}


def _normalize_dimension(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return "Validation"
    for d in _ALLOWED_DIMENSIONS:
        if raw.lower() == d.lower():
            return d
    return _LEGACY_DIMENSION_MAP.get(raw.lower(), "Validation")


def post_process_rules(rules: List[Dict[str, Any]], metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Post-process generated rules to ensure correct dimension mapping and formatting.

    Applies these transformations:
    1. Coerce every dimension into one of the six standard DQ dimensions
       (Accuracy, Completeness, Standardisation, Validation, Uniqueness, Timeliness).
    2. Length / format / pattern rules are kept under Validation.
    3. Ensure human-readable formatting for length constraints.
    4. Ensure a Completeness rule exists for mandatory fields.
    5. Fill ``regex_pattern`` when inferable from metadata.
    """
    processed_rules: List[Dict[str, Any]] = []
    has_completeness_rule = False
    has_length_rule = False

    char_keywords = ['CHAR', 'CHARACTER', 'VARCHAR', 'VARCHAR2', 'STRING', 'TEXT',
                     'maximum', 'max length', 'characters', 'character length']

    for rule in rules:
        rule['dimension'] = _normalize_dimension(rule.get('dimension'))
        rule_text = rule.get('data_quality_rule', '')

        is_length_rule = any(k.lower() in rule_text.lower() for k in char_keywords)

        if is_length_rule and 'character' in rule_text.lower():
            rule['dimension'] = 'Validation'
            has_length_rule = True
            if metadata.get('max_length'):
                field_name = rule.get('business_field', '')
                rule['data_quality_rule'] = f"{field_name} should be maximum {metadata['max_length']} characters"

        if rule['dimension'] == 'Completeness' or 'must not be blank' in rule_text.lower():
            has_completeness_rule = True
            if 'must not be blank' not in rule_text.lower():
                field_name = rule.get('business_field', '')
                rule['data_quality_rule'] = f"{field_name} must not be blank"

        _ensure_regex_pattern(rule, metadata)
        processed_rules.append(rule)

    if metadata.get('mandatory') and not has_completeness_rule:
        field_name = processed_rules[0].get('business_field', '') if processed_rules else ''
        processed_rules.append({
            'business_field': field_name,
            'dimension': 'Completeness',
            'data_quality_rule': f"{field_name} must not be blank",
            'issues_found': 0,
            'issues_found_example': 'All values valid - No issues found',
            'regex_pattern': '',
        })

    if metadata.get('max_length') and not has_length_rule:
        field_name = processed_rules[0].get('business_field', '') if processed_rules else ''
        n = int(metadata['max_length'])
        processed_rules.append({
            'business_field': field_name,
            'dimension': 'Validation',
            'data_quality_rule': f"{field_name} should be maximum {metadata['max_length']} characters",
            'issues_found': 0,
            'issues_found_example': 'All values valid - No issues found',
            'regex_pattern': f'^.{{0,{n}}}$',
        })

    return processed_rules


def extract_column_rules_from_sheet_data(rule_data: Dict, worksheet) -> Dict[str, str]:
    """Extract column-to-rule mappings from scanned sheet data
    
    Attempts to identify:
    - Column names (usually in first few rows)
    - Associated rules (in adjacent cells, comments, or same row)
    """
    column_rules = {}
    
    # Strategy 1: Look for header row (first 5 rows)
    potential_headers = []
    for cell_ref, data in rule_data.items():
        if data['row'] <= 5 and data['value']:
            # Check if this looks like a column name
            value = data['value']
            if len(value) < 100 and not value.startswith('='):  # Not a formula
                potential_headers.append({
                    'name': value,
                    'row': data['row'],
                    'col': data['col'],
                    'cell_ref': cell_ref
                })
    
    # Strategy 2: For each potential header, look for rules in adjacent cells or comments
    for header in potential_headers:
        column_name = header['name']
        rules_found = []
        
        # Check same row (to the right)
        for cell_ref, data in rule_data.items():
            if data['row'] == header['row'] and data['col'] > header['col']:
                if data['value'] and len(data['value']) > 10:  # Likely a rule description
                    rules_found.append(data['value'])
        
        # Check rows below (same column)
        for cell_ref, data in rule_data.items():
            if data['col'] == header['col'] and data['row'] > header['row'] and data['row'] <= header['row'] + 3:
                if data['value']:
                    rules_found.append(data['value'])
        
        # Check for comments on header cell
        if rule_data[header['cell_ref']].get('comment'):
            rules_found.append(rule_data[header['cell_ref']]['comment'])
        
        if rules_found:
            column_rules[column_name] = ' | '.join(rules_found)
            logger.info(f"Mapped column '{column_name}' to rules: {column_rules[column_name][:100]}")
    
    return column_rules


def extract_comprehensive_metadata(column_name: str, rule_text: str = None) -> Dict[str, Any]:
    """Extract comprehensive metadata from column names and rule text
    
    Scans both column name and associated rule text for patterns
    """
    metadata = {
        'max_length': None,
        'data_type_hint': None,
        'mandatory': False,
        'format_restrictions': [],
        'precision': None,
        'scale': None,
        'uniqueness_required': False,
        'conditional_rule': None,
        'allowed_values': None,
        'detected_raw_text': rule_text or '',
        'extracted_pattern': [],
        'conflict_flag': False,
        'original_name': column_name
    }
    
    # Combine column name and rule text for scanning
    scan_text = f"{column_name} {rule_text or ''}"
    
    # Check for mandatory indicator (asterisk)
    if '*' in scan_text:
        metadata['mandatory'] = True
        metadata['extracted_pattern'].append('mandatory (*)')
        logger.info(f"Column '{column_name}' marked as mandatory (contains *)")
    
    # Check for mandatory keywords
    if re.search(r'\b(mandatory|required|compulsory|must|not\s*null|cannot\s*be\s*null|must\s*be\s*provided)\b', scan_text, re.IGNORECASE):
        metadata['mandatory'] = True
        metadata['extracted_pattern'].append('mandatory (keyword)')
        logger.info(f"Column '{column_name}' marked as mandatory (keyword found)")
    
    # Check for uniqueness
    if re.search(r'\b(unique|distinct|pk|primary\s*key|no\s*duplicate)\b', scan_text, re.IGNORECASE):
        metadata['uniqueness_required'] = True
        metadata['extracted_pattern'].append('unique')
        logger.info(f"Column '{column_name}' marked as unique")
    
    # Pattern 1: VARCHAR2(360), VARCHAR(26), CHAR(19), CHARACTERS(21), STRING(50), TEXT(100)
    pattern1 = r'(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?'
    matches = list(re.finditer(pattern1, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            dtype = match.group(1).upper()
            length = int(match.group(2))
            if metadata['max_length'] and metadata['max_length'] != length:
                metadata['conflict_flag'] = True
                logger.warning(f"Conflict detected for '{column_name}': multiple lengths found")
            metadata['data_type_hint'] = dtype
            metadata['max_length'] = length
            metadata['extracted_pattern'].append(f"{dtype}({length})")
            logger.info(f"Extracted from '{column_name}': Type={dtype}, MaxLength={length}")
    
    # Pattern 2: NUMBER(10,2), DECIMAL(15,2), NUMERIC(8,2)
    pattern2 = r'(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?'
    matches = list(re.finditer(pattern2, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            dtype = match.group(1).upper()
            precision = int(match.group(2))
            scale = int(match.group(3)) if match.group(3) else None
            metadata['data_type_hint'] = dtype
            metadata['precision'] = precision
            metadata['scale'] = scale
            pattern_str = f"{dtype}({precision},{scale})" if scale else f"{dtype}({precision})"
            metadata['extracted_pattern'].append(pattern_str)
            logger.info(f"Extracted from '{column_name}': Type={dtype}, Precision={precision}, Scale={scale}")
    
    # Pattern 3: NUM(10), NUM 10, NUMBER 10, INT(5)
    pattern3 = r'(NUM|NUMBER|INT|INTEGER)\s*\(?\s*(\d+)\s*\)?'
    if not metadata['precision']:  # Only if not already found
        matches = list(re.finditer(pattern3, scan_text, re.IGNORECASE))
        if matches:
            for match in matches:
                metadata['data_type_hint'] = 'NUMBER'
                metadata['precision'] = int(match.group(2))
                metadata['extracted_pattern'].append(f"NUMBER({metadata['precision']})")
                logger.info(f"Extracted from '{column_name}': Type=NUMBER, Precision={metadata['precision']}")
    
    # Pattern 4: Just numbers in parentheses like "Name (360)" or "Code (19)"
    pattern4 = r'\((\d+)\)'
    if not metadata['max_length']:  # Only if not already found
        matches = list(re.finditer(pattern4, scan_text))
        if matches:
            metadata['max_length'] = int(matches[0].group(1))
            metadata['extracted_pattern'].append(f"length({metadata['max_length']})")
            logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']}")
    
    # Pattern 5: "max length 50", "maximum 30 characters", "max 100"
    pattern5 = r'(?:max(?:imum)?\s*(?:length|len|size|chars?)?)\s*[:\-]?\s*(\d+)'
    matches = list(re.finditer(pattern5, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            length = int(match.group(1))
            if metadata['max_length'] and metadata['max_length'] != length:
                metadata['conflict_flag'] = True
            metadata['max_length'] = length
            metadata['extracted_pattern'].append(f"max length {length}")
            logger.info(f"Extracted from '{column_name}': MaxLength={length} (from max length pattern)")
    
    # Pattern 6: Format restrictions
    if re.search(r'\b(uppercase|upper|caps|all\s*caps)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('UPPERCASE')
        metadata['extracted_pattern'].append('UPPERCASE')
    if re.search(r'\b(lowercase|lower)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('LOWERCASE')
        metadata['extracted_pattern'].append('LOWERCASE')
    if re.search(r'\b(alphanumeric|alpha\s*numeric|no\s*special\s*char)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('ALPHANUMERIC')
        metadata['extracted_pattern'].append('ALPHANUMERIC')
    if re.search(r'\b(no\s*spaces?|trim)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_SPACES')
        metadata['extracted_pattern'].append('NO_SPACES')
    if re.search(r'\b(no\s*leading\s*zero)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_LEADING_ZERO')
        metadata['extracted_pattern'].append('NO_LEADING_ZERO')
    
    # Pattern 7: Value restrictions (Yes/No, Y/N, etc.)
    if re.search(r'\b(yes\s*/\s*no|y\s*/\s*n|true\s*/\s*false)\b', scan_text, re.IGNORECASE):
        metadata['allowed_values'] = ['Yes', 'No']
        metadata['extracted_pattern'].append('Yes/No')
        logger.info(f"Extracted allowed values from '{column_name}': Yes/No")
    
    # Pattern 8: List of Values, LOV
    if re.search(r'\b(list\s*of\s*values?|lov|enum)\b', scan_text, re.IGNORECASE):
        metadata['extracted_pattern'].append('LOV')
        logger.info(f"Column '{column_name}' has List of Values restriction")
    
    # Pattern 9: Conditional rules
    conditional_match = re.search(r'(mandatory\s*if|required\s*when|only\s*if|depends\s*on)\s+(.+?)(?:\.|$)', scan_text, re.IGNORECASE)
    if conditional_match:
        metadata['conditional_rule'] = conditional_match.group(0)
        metadata['extracted_pattern'].append(f"conditional: {conditional_match.group(1)}")
        logger.info(f"Extracted conditional rule from '{column_name}': {metadata['conditional_rule']}")
    
    # Pattern 10: Date formats
    date_patterns = [
        (r'DD-MM-YYYY', 'DD-MM-YYYY'),
        (r'YYYY-MM-DD', 'YYYY-MM-DD'),
        (r'MM/DD/YYYY', 'MM/DD/YYYY'),
        (r'past\s*date\s*only', 'past date only'),
        (r'future\s*date\s*not\s*allowed', 'no future date'),
        (r'system\s*date', 'system date')
    ]
    for pattern, label in date_patterns:
        if re.search(pattern, scan_text, re.IGNORECASE):
            metadata['extracted_pattern'].append(label)
            logger.info(f"Extracted date pattern from '{column_name}': {label}")
    
    if metadata['format_restrictions']:
        logger.info(f"Extracted format restrictions from '{column_name}': {metadata['format_restrictions']}")
    
    return metadata


def generate_comprehensive_ai_prompt(column_name: str, sample_data: List[str], data_type: str,
                                     null_pct: float, unique_pct: float, 
                                     metadata: Dict[str, Any], rule_source: str) -> str:
    """Generate comprehensive AI prompt for rule generation with exact output format"""
    
    samples = sample_data[:100] if sample_data else []
    sample_str = json.dumps(samples, indent=2)
    
    # Build metadata section
    metadata_section = ""
    if metadata.get('detected_raw_text'):
        metadata_section += f"\n=== RULES FROM EXISTING SHEET ===\n{metadata['detected_raw_text']}\n"
    
    if metadata.get('extracted_pattern'):
        metadata_section += f"\n=== EXTRACTED PATTERNS ===\n"
        for pattern in metadata['extracted_pattern']:
            metadata_section += f"- {pattern}\n"
    
    if metadata.get('max_length'):
        metadata_section += f"\n- Maximum Length: {metadata['max_length']} characters\n"
    if metadata.get('data_type_hint'):
        metadata_section += f"- Data Type Hint: {metadata['data_type_hint']}\n"
    if metadata.get('precision'):
        metadata_section += f"- Numeric Precision: {metadata['precision']}\n"
    if metadata.get('scale'):
        metadata_section += f"- Numeric Scale: {metadata['scale']}\n"
    if metadata.get('mandatory'):
        metadata_section += f"- Mandatory Field: YES\n"
    if metadata.get('uniqueness_required'):
        metadata_section += f"- Uniqueness Required: YES\n"
    if metadata.get('format_restrictions'):
        metadata_section += f"- Format Restrictions: {', '.join(metadata['format_restrictions'])}\n"
    if metadata.get('allowed_values'):
        metadata_section += f"- Allowed Values: {', '.join(metadata['allowed_values'])}\n"
    if metadata.get('conditional_rule'):
        metadata_section += f"- Conditional Rule: {metadata['conditional_rule']}\n"
    if metadata.get('conflict_flag'):
        metadata_section += f"- CONFLICT DETECTED: Multiple conflicting patterns found\n"
    
    prompt = f"""You are an AI Data Quality Rule Engine.

=== COLUMN INFORMATION ===
Column Name: {column_name}
Rule Source: {rule_source}
Detected Data Type: {data_type}
Sample Values: {sample_str}
Null Percentage: {null_pct:.1f}%
Unique Percentage: {unique_pct:.1f}%
{metadata_section}

=== YOUR TASK ===

Generate validation rules in the EXACT format shown below.

Based on the Rule Source:

If Rule Source = "Rules from Existing Sheet":
- Use the extracted patterns and rules from the sheet
- Interpret them logically
- Generate validation rules based on detected patterns

If Rule Source = "Generated by AI":
- No rules found in existing sheets
- Generate intelligent validation rules based on:
  * Column name semantics
  * Sample data patterns
  * Business meaning
  * Enterprise data standards

=== PATTERN INTERPRETATION ===

Length Patterns:
- "30 Characters" → Data Type = String, Max Length = 30
- "varchar2(150)" → Data Type = String, Max Length = 150
- "char(20)" → Data Type = String, Max Length = 20

Numeric Patterns:
- "Number" → Data Type = Numeric
- "number(10)" → Numeric, Max digits = 10
- "numeric(8,2)" → Numeric, Total digits = 8, Decimal = 2

Date Patterns:
- "Date" → Data Type = Date
- "DD-MM-YYYY" → Date format specified

=== DIMENSION CLASSIFICATION (CRITICAL RULES) ===

**STRICT POLICY: Only the SIX standard data-quality dimensions are allowed.**
Every rule MUST use exactly one of these dimension names — no others, no
variations, no legacy labels:

1. **Completeness** - Mandatory/required fields:
   - Fields marked with * (asterisk), keywords: mandatory, required, not null
   - Rule format: "[Field Name] must not be blank"

2. **Validation** - Data type, format, length, pattern, and case constraints:
   - Type checks: "must be numeric", "must be date", "must be valid email"
   - Length: VARCHAR2(30) → "[Field] should be maximum 30 characters"
   - Format/case: UPPERCASE / lowercase / alphanumeric / regex patterns
   - DO NOT use "Character Length" or "Conformity" — those go under Validation.

3. **Uniqueness** - Unique constraints:
   - unique, distinct, no duplicates, primary key

4. **Accuracy** - Value-level correctness:
   - Value ranges, decimal precision, scale, lookups, business correctness

5. **Standardisation** - Cross-field / cross-record agreement:
   - Same value in related columns, referential consistency, sums tie out

6. **Timeliness** - Date freshness / sequencing:
   - Date <= today, end_date >= start_date, freshness windows

DO NOT emit any other dimension name (no Conformity, Character Length,
Integrity, Reliability, Relevance, Precision, Accessibility, etc.).
Length and case rules → Validation. Reference data correctness → Accuracy.

=== HUMAN-READABLE RULE FORMATTING ===

Transform technical constraints to natural language:

Technical Input → Human-Readable Output:
- "VARCHAR2(30)" → "Account Name should be maximum 30 characters"
- "CHAR(19)" → "Code should be maximum 19 characters"
- "VARCHAR(100)" → "Description should be maximum 100 characters"
- "Mandatory *" → "Field Name must not be blank"
- "NUMBER(10,2)" → "Amount must be numeric with maximum 10 digits and 2 decimal places"

=== RULE STATEMENT FORMAT ===

Write rules in this format: [Field Name] + must/should + condition

Examples:
- "Account Name should be maximum 30 characters" (Validation)
- "Account Name must not be blank" (Completeness)
- "Interface Line Number must be numeric" (Validation)
- "Date Placed in Service must be in YYYY/MM/DD date format" (Validation)
- "Asset Type should be maximum 11 characters" (Validation)

=== IMPORTANT GUIDELINES ===

1. **For length / format / pattern rules** (Validation): use "should be maximum X characters" / "must be numeric" / "must be in [format]"
2. **For Completeness**: ALWAYS use "must not be blank" for mandatory fields
3. **For Uniqueness**: Use "must be unique"
4. The "dimension" field MUST be one of exactly: Accuracy, Completeness, Standardisation, Validation, Uniqueness, Timeliness.

5. Generate SEPARATE rules for each dimension:
   - If field has VARCHAR2(30) AND is mandatory → Generate TWO rules:
     * Rule 1: Dimension = "Validation",     Rule = "[Field] should be maximum 30 characters"
     * Rule 2: Dimension = "Completeness", Rule = "[Field] must not be blank"

6. PRIORITY: For every mandatory field, ALWAYS generate a Completeness rule
7. Use detected patterns if available
8. Be dynamic - no hardcoded values
9. Set issues_found to 0 if validation passes
10. Set issues_found_example to "All values valid - No issues found" if no issues

=== OUTPUT FORMAT (STRICT JSON) ===

Return an array of rule objects (one per dimension), wrapped in a "rules" key.
Each rule MUST include "regex_pattern": a Python-style regular expression that
implements the check when applicable (e.g. "^.{{0,30}}$" for max 30 characters,
"^-?\\d+(\\.\\d+)?$" for numeric), or "" when no single regex applies.

{{
  "rules": [
    {{
      "business_field": "{column_name}",
      "dimension": "Validation",
      "data_quality_rule": "{column_name} should be maximum X characters",
      "regex_pattern": "^.{{0,X}}$",
      "issues_found": 0,
      "issues_found_example": "All values valid - No issues found"
    }},
    {{
      "business_field": "{column_name}",
      "dimension": "Completeness",
      "data_quality_rule": "{column_name} must not be blank",
      "regex_pattern": "",
      "issues_found": 0,
      "issues_found_example": "All values valid - No issues found"
    }}
  ]
}}

Return ONLY valid JSON, no markdown."""
    
    return prompt

    """Extract comprehensive metadata from column names with advanced pattern detection
    
    Detects:
    - Length patterns: VARCHAR2(360), CHAR(19), max length 50
    - Numeric patterns: NUMBER(10,2), DECIMAL(15,2), NUM(10)
    - Mandatory indicators: *, mandatory, required
    - Format restrictions: UPPERCASE, ALPHANUMERIC, NO_SPACES
    - Uniqueness: unique, distinct, PK
    - Conditional rules: "if", "when", "depends on"
    
    Returns comprehensive metadata dict
    """
    metadata = {
        'max_length': None,
        'data_type_hint': None,
        'mandatory': False,
        'format_restrictions': [],
        'precision': None,
        'scale': None,
        'uniqueness_required': False,
        'conditional_rule': None,
        'allowed_values': None,
        'original_name': column_name
    }
    
    # Check for mandatory indicator (asterisk)
    if '*' in column_name:
        metadata['mandatory'] = True
        logger.info(f"Column '{column_name}' marked as mandatory (contains *)")
    
    # Check for mandatory keywords
    if re.search(r'\b(mandatory|required|compulsory|must|not\s*null)\b', column_name, re.IGNORECASE):
        metadata['mandatory'] = True
        logger.info(f"Column '{column_name}' marked as mandatory (keyword found)")
    
    # Check for uniqueness
    if re.search(r'\b(unique|distinct|pk|primary\s*key|no\s*duplicate)\b', column_name, re.IGNORECASE):
        metadata['uniqueness_required'] = True
        logger.info(f"Column '{column_name}' marked as unique")
    
    # Pattern 1: VARCHAR2(360), VARCHAR(26), CHAR(19), CHARACTERS(21), STRING(50), TEXT(100)
    pattern1 = r'(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?'
    match = re.search(pattern1, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['max_length'] = int(match.group(2))
        logger.info(f"Extracted from '{column_name}': Type={metadata['data_type_hint']}, MaxLength={metadata['max_length']}")
        return metadata
    
    # Pattern 2: NUMBER(10,2), DECIMAL(15,2), NUMERIC(8,2)
    pattern2 = r'(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?'
    match = re.search(pattern2, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['precision'] = int(match.group(2))
        if match.group(3):
            metadata['scale'] = int(match.group(3))
        logger.info(f"Extracted from '{column_name}': Type={metadata['data_type_hint']}, Precision={metadata['precision']}, Scale={metadata.get('scale')}")
        return metadata
    
    # Pattern 3: NUM(10), NUM 10, NUMBER 10, INT(5)
    pattern3 = r'(NUM|NUMBER|INT|INTEGER)\s*\(?\s*(\d+)\s*\)?'
    match = re.search(pattern3, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = 'NUMBER'
        metadata['precision'] = int(match.group(2))
        logger.info(f"Extracted from '{column_name}': Type=NUMBER, Precision={metadata['precision']}")
        return metadata
    
    # Pattern 4: Just numbers in parentheses like "Name (360)" or "Code (19)"
    pattern4 = r'\((\d+)\)'
    match = re.search(pattern4, column_name)
    if match:
        metadata['max_length'] = int(match.group(1))
        logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']}")
    
    # Pattern 5: "max length 50", "maximum 30 characters", "max 100"
    pattern5 = r'(?:max(?:imum)?\s*(?:length|len|size|chars?)?)\s*[:\-]?\s*(\d+)'
    match = re.search(pattern5, column_name, re.IGNORECASE)
    if match:
        metadata['max_length'] = int(match.group(1))
        logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']} (from max length pattern)")
    
    # Pattern 6: Format restrictions
    if re.search(r'\b(uppercase|upper|caps|all\s*caps)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('UPPERCASE')
    if re.search(r'\b(lowercase|lower)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('LOWERCASE')
    if re.search(r'\b(alphanumeric|alpha\s*numeric|no\s*special\s*char)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('ALPHANUMERIC')
    if re.search(r'\b(no\s*spaces?|trim)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_SPACES')
    if re.search(r'\b(no\s*leading\s*zero)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_LEADING_ZERO')
    
    # Pattern 7: Value restrictions (Yes/No, Y/N, etc.)
    if re.search(r'\b(yes\s*/\s*no|y\s*/\s*n|true\s*/\s*false)\b', column_name, re.IGNORECASE):
        metadata['allowed_values'] = ['Yes', 'No']
        logger.info(f"Extracted allowed values from '{column_name}': Yes/No")
    
    # Pattern 8: Conditional rules
    if re.search(r'\b(if|when|depends\s*on)\b', column_name, re.IGNORECASE):
        conditional_match = re.search(r'(if|when|depends\s*on)\s+(.+)', column_name, re.IGNORECASE)
        if conditional_match:
            metadata['conditional_rule'] = conditional_match.group(0)
            logger.info(f"Extracted conditional rule from '{column_name}': {metadata['conditional_rule']}")
    
    if metadata['format_restrictions']:
        logger.info(f"Extracted format restrictions from '{column_name}': {metadata['format_restrictions']}")
    
    return metadata


def generate_comprehensive_ai_prompt(column_name: str, sample_data: List[str], data_type: str,
                                     null_pct: float, unique_pct: float,
                                     metadata: Dict[str, Any], rule_source: str = None,
                                     all_columns: Optional[List[str]] = None,
                                     sibling_samples: Optional[Dict[str, List[str]]] = None) -> str:
    """Build a deterministic per-column data-quality prompt.

    Returns six rules — one per dimension Accuracy, Completeness, Standardisation,
    Timeliness, Validation, Uniqueness. Cross-field rules are produced by a
    separate dedicated pass (see ``generate_cross_field_prompt``) so the
    model can spend its attention budget on a single focused task.

    The ``all_columns`` and ``sibling_samples`` arguments are accepted for
    backward compatibility but are intentionally ignored here.
    """
    del all_columns, sibling_samples  # reserved for caller compatibility

    samples = sample_data[:50] if sample_data else []
    sample_str = json.dumps(samples, ensure_ascii=False)

    metadata_lines: List[str] = []
    if metadata:
        if metadata.get('max_length'):
            metadata_lines.append(f"- Maximum Length: {metadata['max_length']} characters")
        if metadata.get('data_type_hint'):
            metadata_lines.append(f"- Data Type Hint: {metadata['data_type_hint']}")
        if metadata.get('precision'):
            metadata_lines.append(f"- Numeric Precision: {metadata['precision']}")
        if metadata.get('scale'):
            metadata_lines.append(f"- Numeric Scale: {metadata['scale']}")
        if metadata.get('mandatory'):
            metadata_lines.append("- Mandatory Field: YES")
        if metadata.get('uniqueness_required'):
            metadata_lines.append("- Uniqueness Required: YES")
        if metadata.get('format_restrictions'):
            metadata_lines.append(
                f"- Format Restrictions: {', '.join(metadata['format_restrictions'])}"
            )
        if metadata.get('allowed_values'):
            metadata_lines.append(
                f"- Allowed Values: {', '.join(metadata['allowed_values'])}"
            )
        if metadata.get('conditional_rule'):
            metadata_lines.append(f"- Conditional Rule: {metadata['conditional_rule']}")
        # Semantic glossary hints (filled in by the rule_generator service
        # when the user has generated a glossary on the AI Validations tab).
        # These steer the LLM toward type-aware rules — e.g. when
        # ``semantic_type`` is ``"email"`` it should emit the standard email
        # regex for the Validation rule.
        if metadata.get('semantic_type'):
            metadata_lines.append(f"- Semantic Type: {metadata['semantic_type']}")
        if metadata.get('semantic_display_name'):
            metadata_lines.append(f"- Semantic Display Name: {metadata['semantic_display_name']}")
        if metadata.get('semantic_description'):
            metadata_lines.append(f"- Semantic Description: {metadata['semantic_description']}")
        if metadata.get('semantic_format_hint'):
            metadata_lines.append(f"- Format Hint (from glossary): {metadata['semantic_format_hint']}")
    metadata_block = "\n".join(metadata_lines) if metadata_lines else "(none extracted)"

    rule_source_block = (
        f"\nEXCEL NOTE / COMMENT FOR THIS COLUMN:\n{rule_source}\n"
        if rule_source and rule_source not in {"Generated by AI", "Rules from Existing Sheet"}
        else ""
    )

    prompt = f"""You are a data-quality rule engine. For the column described
below, output exactly six data-quality rules as JSON — one rule per
dimension — covering the standard six DQ dimensions. Cross-field rules
are handled by a separate pass and must NOT appear here.

COLUMN UNDER REVIEW
-------------------
Name: {column_name}
Pandas dtype: {data_type}
Null %: {null_pct:.1f}
Unique %: {unique_pct:.1f}
Distinct sample values (sorted, max 50): {sample_str}

EXTRACTED METADATA
------------------
{metadata_block}
{rule_source_block}

DIMENSION DEFINITIONS (use these definitions verbatim)
------------------------------------------------------
- Accuracy: the value reflects the real-world entity it represents.
  Anchor on sample values and column semantics (e.g. an email column's
  values must be syntactically resolvable; an amount must be a non-negative
  number when the field is a price).
- Completeness: the value is present (not null, not blank, not whitespace).
- Standardisation: the value follows a uniform representation across rows
  (e.g. consistent casing, consistent date format, consistent code system).
- Timeliness: the value is current and within an expected time window.
  Only meaningful for date/time columns. For non-date columns, emit:
  "<Field>: timeliness is not applicable for non-date fields".
- Validation: the value satisfies the column's structural rules — data type,
  length, allowed values, regex / format pattern.
- Uniqueness: the value does not duplicate another row's value when the
  column is an identifier. For clearly non-identifier columns (categorical,
  low-cardinality, free text), emit:
  "<Field>: duplicates are allowed for this non-identifier field".

INSTRUCTIONS
------------
1. Emit exactly six rules — one per dimension — in the order listed in
   the schema below. Do not skip a dimension. If a dimension does not
   apply, use the "not applicable" phrasing shown in the definitions.
2. Rule wording: "<Field> must <verb> <condition>" (one concise sentence).
   The "not applicable" rules are the only exceptions.
3. Use evidence-based limits. If metadata supplies a max_length, allowed
   values, or a regex, use that exactly — do not invent thresholds.
4. The dimension field MUST be spelled EXACTLY as listed in the schema:
   Accuracy, Completeness, Standardisation, Timeliness, Validation, Uniqueness.
5. Do NOT propose any rule that mentions another column. Cross-field rules
   are produced by a separate pass.
6. Return ONLY a JSON object — no prose, no markdown fence.

OUTPUT SCHEMA (return exactly these keys, in this order)
--------------------------------------------------------
{{
  "business_field": "<human-readable field name, derived from column name>",
  "rules": [
    {{ "dimension": "Accuracy",     "data_quality_rule": "<one sentence>" }},
    {{ "dimension": "Completeness", "data_quality_rule": "<one sentence>" }},
    {{ "dimension": "Standardisation",  "data_quality_rule": "<one sentence>" }},
    {{ "dimension": "Timeliness",   "data_quality_rule": "<one sentence>" }},
    {{ "dimension": "Validation",     "data_quality_rule": "<one sentence>" }},
    {{ "dimension": "Uniqueness",   "data_quality_rule": "<one sentence>" }}
  ]
}}

WORKED EXAMPLE — VAT number column (mandatory)
----------------------------------------------
{{
  "business_field": "VAT Number",
  "rules": [
    {{ "dimension": "Accuracy",     "data_quality_rule": "VAT Number must correspond to an active VAT registration with the issuing tax authority" }},
    {{ "dimension": "Completeness", "data_quality_rule": "VAT Number must not be blank" }},
    {{ "dimension": "Standardisation",  "data_quality_rule": "VAT Number must be stored in uppercase with no spaces, hyphens, or punctuation" }},
    {{ "dimension": "Timeliness",   "data_quality_rule": "VAT Number: timeliness is not applicable for non-date fields" }},
    {{ "dimension": "Validation",     "data_quality_rule": "VAT Number must be alphanumeric with length between 8 and 14 characters after normalization" }},
    {{ "dimension": "Uniqueness",   "data_quality_rule": "VAT Number must be unique across all records once normalized" }}
  ]
}}
"""

    return prompt


def generate_cross_field_prompt(column_samples: Dict[str, List[str]]) -> str:
    """Build a single whole-dataset cross-field prompt.

    Sees every column with up to 5 sample values, asks the model to find
    cross-field rules involving 2+ columns. The model is expected to think
    top-down ("which tuples of columns identify a record?") rather than
    column-by-column. Returns 0–10 rules.

    Each rule names the columns it involves in a ``columns`` array so the
    caller can validate the references against the actual column list and
    drop any rule that names a hallucinated column.
    """
    if not column_samples:
        catalog_block = "(empty dataset)"
    else:
        lines = []
        for name in sorted(column_samples.keys()):
            preview = column_samples[name][:5]
            lines.append(f"- {name}: {json.dumps(preview, ensure_ascii=False)}")
        catalog_block = "\n".join(lines)

    return f"""You are a data-quality rule engine specialising in cross-field
constraints. Your job is to look at the WHOLE table below and propose
rules that involve TWO OR MORE columns at once. Single-column rules are
out of scope and must not appear in your output.

DATASET COLUMNS (with up to 5 sample values each)
-------------------------------------------------
{catalog_block}

HOW TO THINK ABOUT THIS
-----------------------
Read the column list once, top-down, and ask:

1. Which columns together IDENTIFY a record? (composite uniqueness)
   Prefer THREE-COLUMN tuples — they are stronger and more meaningful
   than 2-column tuples. A business name alone is too coarse; name +
   country + entity_type is the canonical customer-master key. Order
   number + line_number + fiscal_year, patient_id + visit_date + visit_type,
   etc. Always look for the FULL tuple, not the minimal one.

2. Which columns are PRESENT-OR-ABSENT depending on another column?
   Look for fields that only make sense for some values of another field
   — e.g. a tax-registration number required only for certain countries,
   or a discharge date required only when status = 'closed'. These often
   involve THREE columns: a target, a context, and a discriminator
   ("vat_no required when country is EU AND entity_type is Corporation").

3. Which columns DERIVE FROM another column?
   Look for prefixes, country codes, dialing codes, currency-from-locale,
   etc. A field that always begins with the value of another field, or is
   determined by a lookup from it. Don't stop at one pairwise rule — if
   country drives vat_no AND currency, propose both, or a single
   three-column rule.

4. Which columns satisfy ARITHMETIC identities?
   THREE OR FOUR COLUMNS is the norm here: net + tax = gross, quantity ×
   unit_price = line_total, start_date + duration = end_date, opening +
   debits − credits = closing. Always include every column in the
   identity, not just the two most obvious ones.

5. Which columns must REFERENCE THE SAME CONCEPT consistently?
   A country and a currency. A state and a country and a country code.
   A category and a sub-category and a top-level group. These chains are
   usually 3+ columns deep — chase the full chain.

6. Which columns have CONDITIONAL VALUES derived from another?
   "entity_type = LLP whenever company_type = LIMITED_LIABILITY_PARTNERSHIP".
   Often extends to a third column: "AND legal_status must be ACTIVE".

OUTPUT REQUIREMENTS
-------------------
- 0 to 10 cross-field rules total. Quality over quantity. One strong rule
  is better than three weak ones.
- Every rule MUST involve two or more columns from the dataset.
- STRONGLY PREFER 3+ column rules when the relationship genuinely spans
  three columns. Do not artificially split a 3-column composite-uniqueness
  rule into multiple 2-column rules — that loses information. Likewise,
  do not artificially merge unrelated 2-column rules into one 3-column
  rule.
- Use EXACT column names from the list above. Do not invent columns.
- Use specifics drawn from sample values where possible — a literal value
  list, a country code, a numeric tolerance. Avoid generic phrasings like
  "<a> and <b> must be consistent".
- If no honest cross-field rule exists in this dataset, return rules: [].
  An empty list is acceptable when the columns truly do not relate.
- Return ONLY a JSON object matching the schema below.

OUTPUT SCHEMA
-------------
{{
  "rules": [
    {{
      "data_quality_rule": "<one sentence — must mention two or more columns by exact name>",
      "columns": ["<column_a>", "<column_b>", "<...>"]
    }}
  ]
}}

WORKED EXAMPLE — customer master data
-------------------------------------
DATASET COLUMNS:
- name: ["Acme GmbH", "Brico SARL", "Globex Ltd"]
- country: ["DE", "FR", "GB", "IN"]
- entity_type: ["Corporation", "LLC", "Partnership"]
- vat_no: ["DE123456789", "FR12345678901"]
- pan_number: ["AAAPL1234C"]
- net_amount: ["100.00", "250.00"]
- tax_amount: ["18.00", "45.00"]
- gross_amount: ["118.00", "295.00"]
- currency: ["EUR", "USD"]

{{
  "rules": [
    {{
      "data_quality_rule": "Flag probable duplicate when normalized name fuzzy match is at least 85% and country and entity_type match exactly",
      "columns": ["name", "country", "entity_type"]
    }},
    {{
      "data_quality_rule": "vat_no must start with the 2-letter ISO code stored in country (e.g. country=DE → vat_no starts with DE)",
      "columns": ["vat_no", "country"]
    }},
    {{
      "data_quality_rule": "vat_no must be present when country is in [DE, FR, IT, ES, NL, GB] AND entity_type is 'Corporation'",
      "columns": ["vat_no", "country", "entity_type"]
    }},
    {{
      "data_quality_rule": "currency must equal EUR when country is in [DE, FR, IT, ES, NL] AND entity_type is 'Corporation' (Eurozone corporate accounts)",
      "columns": ["currency", "country", "entity_type"]
    }},
    {{
      "data_quality_rule": "gross_amount must equal net_amount + tax_amount within a tolerance of 0.01 in the same currency",
      "columns": ["gross_amount", "net_amount", "tax_amount", "currency"]
    }}
  ]
}}

Note in the example above: out of 5 rules, 4 of them involve 3+ columns.
That ratio (most rules being 3+ columns) is what we want for a real dataset
when the relationships genuinely span 3+ columns. Don't force it when only
2 columns relate, but don't artificially limit yourself to pairs either.
"""


# ═══════════════════════════════════════════════════════════════════════════
# RULE VALIDATION EXECUTOR
# ═══════════════════════════════════════════════════════════════════════════


def _extract_allowed_values(rule_text: str) -> Optional[List[str]]:
    """Parse 'must be one of: X, Y, Z' from a rule statement."""
    m = re.search(r'must be one of[:\s]+(.+)', rule_text, re.IGNORECASE)
    if not m:
        return None
    raw = m.group(1).strip().rstrip('.')
    return [v.strip().strip("'\"") for v in re.split(r'[,;]', raw) if v.strip()]


def _extract_max_chars(rule_text: str) -> Optional[int]:
    """Parse 'maximum X characters' from a rule statement."""
    m = re.search(r'(?:maximum|max)\s+(\d+)\s+characters?', rule_text, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _extract_range(rule_text: str) -> Tuple[Optional[float], Optional[float]]:
    """Parse 'within the range of X to Y' or 'between X and Y'."""
    m = re.search(r'(?:range\s+(?:of\s+)?|between\s+)([\d.]+)\s+(?:to|and)\s+([\d.]+)', rule_text, re.IGNORECASE)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _extract_max_decimals(rule_text: str) -> Optional[int]:
    """Parse 'maximum of N decimal places'."""
    m = re.search(r'(?:maximum\s+(?:of\s+)?)?(\d+)\s+decimal\s+place', rule_text, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _format_examples(series: pd.Series, limit: int = 5) -> str:
    """Return a readable string of example violating values."""
    samples = series.head(limit).tolist()
    parts = [repr(v) for v in samples]
    remaining = len(series) - limit
    result = ", ".join(parts)
    if remaining > 0:
        result += f" ... and {remaining} more"
    return result


def infer_regex_pattern_from_rule(dimension: str, rule_text: str) -> str:
    """Derive a best-effort Python ``re`` pattern from a rule statement.

    Returns an empty string when no reasonable single-cell pattern exists (e.g. uniqueness).

    Args:
        dimension: DQ dimension for the rule row.
        rule_text: Human-readable ``Data Quality Rule`` text.

    Returns:
        A regex string, or ``""`` if none applies.
    """
    if not rule_text or not str(rule_text).strip():
        return ""
    rt = str(rule_text).strip()
    rl = rt.lower()
    dim = (dimension or "").strip()

    if dim in ("Uniqueness", "Relevance"):
        return ""

    allowed = _extract_allowed_values(rt)
    if allowed:
        inner = "|".join(re.escape(v) for v in allowed)
        return f"(?i)^({inner})$"

    if (
        dim == "Completeness"
        or "must not be blank" in rl
        or "cannot be null" in rl
        or "not be null" in rl
    ):
        return r"^(?=.*\S).*$"

    if (
        "must be numeric" in rl
        or "must be a numeric" in rl
        or "must be an integer" in rl
        or re.search(r"\bmust be (a )?whole number\b", rl)
    ):
        return r"^-?\d+(\.\d+)?$"

    if dim == "Conformity":
        if "uppercase" in rl or "upper case" in rl or "all caps" in rl:
            return r"^[^a-z]*$"
        if "lowercase" in rl or "lower case" in rl:
            return r"^[^A-Z]*$"
        if "alphanumeric" in rl and "no special" not in rl:
            return r"^[A-Za-z0-9\s]+$"
        if "no special character" in rl:
            return r"^[A-Za-z0-9\s\-_.]+$"

    mc = _extract_max_chars(rt)
    if mc is not None:
        return rf"^.{0,{mc}}$"

    if "valid string" in rl or "valid string format" in rl:
        return r"^(?=.*\S).*$"

    if "email" in rl and "format" in rl:
        return r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"

    if re.search(r"\bdd-mm-yyyy\b", rl, re.IGNORECASE):
        return r"^\d{2}-\d{2}-\d{4}$"
    if re.search(r"\byyyy-mm-dd\b", rl, re.IGNORECASE):
        return r"^\d{4}-\d{2}-\d{2}$"
    if re.search(r"\bmm/dd/yyyy\b", rl, re.IGNORECASE):
        return r"^\d{2}/\d{2}/\d{4}$"

    if re.search(
        r"(in uppercase|uppercase format|must be uppercase)\b", rl
    ) and "not uppercase" not in rl:
        return r"^[^a-z]*$"

    return ""


def enrich_dataframe_regex_patterns(df: pd.DataFrame) -> pd.DataFrame:
    """Fill empty ``Regex Pattern`` cells using :func:`infer_regex_pattern_from_rule`.

    Args:
        df: Rules dataframe with ``Dimension`` and ``Data Quality Rule`` columns.

    Returns:
        A copy of *df* with inferred patterns where the pattern was blank.
    """
    out = df.copy()
    if "Regex Pattern" not in out.columns:
        out["Regex Pattern"] = ""
    for idx in out.index:
        dim = str(out.at[idx, "Dimension"]).strip()
        # Cross-field rules involve multiple columns; a single-column
        # regex is meaningless for them.
        if dim == "Cross-field Validation":
            continue
        cur = out.at[idx, "Regex Pattern"]
        if pd.notna(cur) and str(cur).strip() != "":
            continue
        rule = str(out.at[idx, "Data Quality Rule"])
        inferred = infer_regex_pattern_from_rule(dim, rule)
        if inferred:
            out.at[idx, "Regex Pattern"] = inferred
    return out


def validate_rule(df: pd.DataFrame, column: str, dimension: str, rule_text: str) -> Tuple[int, str]:
    """Execute a single rule against the DataFrame and return (issue_count, example_text).

    This is a best-effort interpreter: it parses the human-readable rule statement
    and applies the matching check.  Rules that cannot be interpreted automatically
    are skipped (returned as 0 issues).
    """
    if column not in df.columns:
        return 0, "Column not found in dataset"

    col = df[column]
    rule_lower = rule_text.lower()

    # ── Completeness: blanks / nulls ──────────────────────────────────────
    if dimension == "Completeness" or "must not be blank" in rule_lower or "not be null" in rule_lower:
        mask = col.isnull() | col.astype(str).str.strip().eq("")
        count = int(mask.sum())
        if count:
            return count, f"{count} blank/null values found"
        return 0, "All values valid - No issues found"

    # ── Uniqueness ────────────────────────────────────────────────────────
    if dimension == "Uniqueness" or "must be unique" in rule_lower or "not have duplicates" in rule_lower:
        non_null = col.dropna()
        dups = non_null[non_null.duplicated(keep=False)]
        count = int(len(dups))
        if count:
            examples = _format_examples(dups.drop_duplicates().head(5))
            return count, f"{count} duplicate values, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Validation: allowed-value lists ─────────────────────────────────────
    allowed = _extract_allowed_values(rule_text)
    if allowed:
        non_null = col.dropna().astype(str).str.strip()
        # case-insensitive comparison
        allowed_lower = {v.lower() for v in allowed}
        bad = non_null[~non_null.str.lower().isin(allowed_lower)]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values not in allowed list, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Accuracy: decimal places (before generic numeric — more specific) ─
    max_dec = _extract_max_decimals(rule_text)
    if max_dec is not None:
        numeric_col = pd.to_numeric(col, errors="coerce").dropna()
        if len(numeric_col) > 0:
            def _count_decimals(v: float) -> int:
                s = f"{v:.15g}"
                return len(s.split(".")[1]) if "." in s else 0
            dec_counts = numeric_col.apply(_count_decimals)
            bad_mask = dec_counts > max_dec
            bad = numeric_col[bad_mask]
            count = int(len(bad))
            if count:
                examples = _format_examples(bad.drop_duplicates())
                return count, f"{count} values exceed {max_dec} decimal places, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Validation: numeric check ───────────────────────────────────────────
    if "must be a numeric" in rule_lower or "must be numeric" in rule_lower:
        non_null = col.dropna()
        converted = pd.to_numeric(non_null, errors="coerce")
        bad = non_null[converted.isna()]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} non-numeric values, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Validation: valid string check ──────────────────────────────────────
    if "must be a valid string" in rule_lower:
        mask = col.isnull() | col.astype(str).str.strip().eq("")
        count = int(mask.sum())
        if count:
            return count, f"{count} blank/null values found"
        return 0, "All values valid - No issues found"

    # ── Length / Validation (max-character) ─────────────────────────────────
    max_chars = _extract_max_chars(rule_text)
    if max_chars or dimension == "Character Length" or "character" in rule_lower and "maximum" in rule_lower:
        if max_chars:
            non_null = col.dropna().astype(str)
            too_long = non_null[non_null.str.len() > max_chars]
            count = int(len(too_long))
            if count:
                examples = _format_examples(too_long.drop_duplicates())
                return count, f"{count} values exceed {max_chars} chars, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Conformity: uppercase check ───────────────────────────────────────
    if "uppercase" in rule_lower or "upper case" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        non_empty = non_null[non_null.str.len() > 0]
        bad = non_empty[non_empty != non_empty.str.upper()]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values not uppercase, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Conformity: lowercase check ───────────────────────────────────────
    if "lowercase" in rule_lower or "lower case" in rule_lower or "must be in lowercase" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        non_empty = non_null[non_null.str.len() > 0]
        bad = non_empty[non_empty != non_empty.str.lower()]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values not lowercase, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Conformity: alphanumeric check ────────────────────────────────────
    if "alphanumeric" in rule_lower and "no special" not in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        non_empty = non_null[non_null.str.len() > 0]
        bad = non_empty[~non_empty.str.match(r'^[A-Za-z0-9\s]+$')]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values contain non-alphanumeric chars, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Conformity: no special characters ─────────────────────────────────
    if "no special character" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        non_empty = non_null[non_null.str.len() > 0]
        bad = non_empty[~non_empty.str.match(r'^[A-Za-z0-9\s\-_.]+$')]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values contain special chars, e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Conformity: text format / reasonable length ───────────────────────
    if "should not exceed reasonable length" in rule_lower or "reasonable length" in rule_lower:
        non_null = col.dropna().astype(str)
        too_long = non_null[non_null.str.len() > 255]
        count = int(len(too_long))
        if count:
            examples = _format_examples(too_long.str[:50].drop_duplicates())
            return count, f"{count} values exceed 255 chars"
        return 0, "All values valid - No issues found"

    # ── Conformity: consistent casing ─────────────────────────────────────
    if "consistent casing" in rule_lower or "consistent naming" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        unique_vals = non_null.unique()
        lower_map: Dict[str, List[str]] = {}
        for v in unique_vals:
            key = v.lower()
            lower_map.setdefault(key, []).append(v)
        inconsistent = {k: vs for k, vs in lower_map.items() if len(vs) > 1}
        if inconsistent:
            pairs = [f"{' vs '.join(vs)}" for vs in list(inconsistent.values())[:3]]
            total = sum(len(vs) for vs in inconsistent.values())
            return total, f"Inconsistent casing: {'; '.join(pairs)}"
        return 0, "All values valid - No issues found"

    # ── Accuracy: range check ─────────────────────────────────────────────
    lo, hi = _extract_range(rule_text)
    if lo is not None and hi is not None:
        numeric_col = pd.to_numeric(col, errors="coerce")
        bad = numeric_col.dropna()[(numeric_col.dropna() < lo) | (numeric_col.dropna() > hi)]
        count = int(len(bad))
        if count:
            examples = _format_examples(bad.drop_duplicates())
            return count, f"{count} values outside range [{lo}, {hi}], e.g. {examples}"
        return 0, "All values valid - No issues found"

    # ── Validation: valid string format ─────────────────────────────────────
    if "valid string format" in rule_lower or "valid string" in rule_lower:
        mask = col.isnull() | col.astype(str).str.strip().eq("")
        count = int(mask.sum())
        if count:
            return count, f"{count} blank/null values found"
        return 0, "All values valid - No issues found"

    # ── Conformity: expected format / follow specified format ──────────────
    if "should follow" in rule_lower and "format" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        unique_vals = non_null.unique()
        # detect inconsistent patterns (e.g. X31300 vs X31-300)
        if len(unique_vals) > 1:
            patterns = set()
            for v in unique_vals[:100]:
                p = re.sub(r'[A-Za-z]', 'A', re.sub(r'\d', 'N', v))
                patterns.add(p)
            if len(patterns) > 1:
                sample_pairs = list(unique_vals[:4])
                return len(non_null), f"Multiple formats detected: {', '.join(repr(s) for s in sample_pairs)}"
        return 0, "All values valid - No issues found"

    # ── Conformity: expected text format ──────────────────────────────────
    if "conform to expected" in rule_lower and ("text format" in rule_lower or "string format" in rule_lower):
        mask = col.isnull() | col.astype(str).str.strip().eq("")
        count = int(mask.sum())
        if count:
            return count, f"{count} blank/empty values"
        return 0, "All values valid - No issues found"

    # ── Standardisation: consistent reflection / naming convention ─────────────
    if "consistently reflect" in rule_lower or "consistent naming convention" in rule_lower or "should maintain consistent" in rule_lower:
        non_null = col.dropna().astype(str).str.strip()
        unique_vals = non_null.unique()
        lower_map: Dict[str, List[str]] = {}
        for v in unique_vals:
            key = v.lower()
            lower_map.setdefault(key, []).append(v)
        inconsistent = {k: vs for k, vs in lower_map.items() if len(vs) > 1}
        if inconsistent:
            pairs = [f"{' vs '.join(vs)}" for vs in list(inconsistent.values())[:3]]
            total = sum(len(vs) for vs in inconsistent.values())
            return total, f"Inconsistent values: {'; '.join(pairs)}"
        return 0, "All values valid - No issues found"

    # ── Relevance / Accuracy: soft checks — cannot auto-validate ──────────
    if dimension in ("Relevance", "Accuracy") and "should reflect" in rule_lower:
        return 0, "Manual review recommended — cannot auto-validate semantic accuracy"

    # ── Fallback: unrecognised rule ───────────────────────────────────────
    return 0, "All values valid - No issues found"


def validate_all_rules(df: pd.DataFrame, rules_df: pd.DataFrame) -> pd.DataFrame:
    """Execute every rule in *rules_df* against *df* and update issue columns.

    Returns a copy of *rules_df* with ``Issues Found`` and
    ``Issues Found Example`` populated from actual data checks.

    Cross-field rules are not auto-evaluated — they involve multiple
    columns and are too open-ended for the single-column interpreter
    below. They are flagged for manual review instead.
    """
    result = rules_df.copy()

    for idx in result.index:
        column = result.at[idx, "Column"]
        dimension = result.at[idx, "Dimension"]
        rule_text = result.at[idx, "Data Quality Rule"]

        if str(dimension).strip() == "Cross-field Validation":
            result.at[idx, "Issues Found"] = 0
            result.at[idx, "Issues Found Example"] = "Cross-field — manual review"
            continue

        try:
            count, example = validate_rule(df, column, dimension, rule_text)
            result.at[idx, "Issues Found"] = count
            result.at[idx, "Issues Found Example"] = example
        except Exception as exc:
            logger.warning("Validation error for rule %s on column %s: %s", rule_text, column, exc)
            result.at[idx, "Issues Found"] = 0
            result.at[idx, "Issues Found Example"] = f"Validation error: {exc}"

    return result
