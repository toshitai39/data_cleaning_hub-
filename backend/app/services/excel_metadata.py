"""Excel + column-name metadata extractors ported 1:1 from features/profiling/ui.py.

Functions copied verbatim:
- extract_column_metadata (lines 73-114)
- extract_excel_cell_notes (lines 117-172)
- _extract_client_rules_from_excel_metadata (lines 972-1057)
- _parse_constraint_to_rules (lines 1060-1160)
- _find_matching_column (lines 1163-1183)
"""
from __future__ import annotations

import logging
import os
import re
from typing import Any, Dict, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)


def extract_column_metadata(column_name: str) -> Dict[str, Any]:
    """Extract metadata from column names like VARCHAR2(360), CHAR(19), etc."""
    metadata = {
        "max_length": None,
        "data_type_hint": None,
        "original_name": column_name,
    }

    pattern1 = r"(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?"
    match = re.search(pattern1, column_name, re.IGNORECASE)
    if match:
        metadata["data_type_hint"] = match.group(1).upper()
        metadata["max_length"] = int(match.group(2))
        return metadata

    pattern2 = r"(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?"
    match = re.search(pattern2, column_name, re.IGNORECASE)
    if match:
        metadata["data_type_hint"] = match.group(1).upper()
        metadata["max_length"] = int(match.group(2))
        return metadata

    pattern3 = r"\((\d+)\)"
    match = re.search(pattern3, column_name)
    if match:
        metadata["max_length"] = int(match.group(1))
        return metadata

    return metadata


def extract_excel_cell_notes(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, str]:
    """Extract cell notes/comments + 2nd-row metadata from Excel file."""
    if not file_path or not os.path.exists(file_path):
        return {}
    suffix = os.path.splitext(file_path)[1].lower()
    if suffix not in (".xlsx", ".xls", ".xlsm"):
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=False)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        column_notes: Dict[str, str] = {}

        for cell in ws[1]:
            if cell.comment:
                column_name = str(cell.value)
                comment_text = cell.comment.text
                column_notes[column_name] = comment_text

        if ws.max_row >= 2:
            for idx, cell in enumerate(ws[2], start=1):
                header_cell = ws.cell(row=1, column=idx)
                column_name = str(header_cell.value)
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if re.search(
                        r"(VARCHAR|CHAR|NUMBER|DECIMAL|DATE|TIMESTAMP)",
                        cell_value, re.IGNORECASE,
                    ):
                        if column_name not in column_notes:
                            column_notes[column_name] = cell_value
                        else:
                            column_notes[column_name] += f" | {cell_value}"
        wb.close()
        return column_notes
    except ImportError:
        logger.warning("openpyxl not installed.")
        return {}
    except Exception as exc:
        logger.error("Error extracting Excel comments: %s", exc)
        return {}


def _parse_constraint_to_rules(constraint: str, column_name: str, is_mandatory: bool) -> List[Dict]:
    """Parse a constraint string into human-readable rules. Verbatim port."""
    rules: List[Dict] = []
    constraint_lower = constraint.lower().strip()

    char_match = re.match(r"^(\d+)\s*characters?$", constraint_lower)
    if char_match:
        max_chars = char_match.group(1)
        rules.append({
            "dimension": "Character Length",
            "rule_statement": f"{column_name} should be maximum {max_chars} characters",
        })
        if is_mandatory:
            rules.append({
                "dimension": "Completeness",
                "rule_statement": f"{column_name} must not be blank",
            })
        return rules

    if "yyyy" in constraint_lower or "mm" in constraint_lower or "dd" in constraint_lower:
        rules.append({
            "dimension": "Validation",
            "rule_statement": f"{column_name} must be in {constraint} date format",
        })
        rules.append({
            "dimension": "Timeliness",
            "rule_statement": f"{column_name} should not be future dated",
        })
        if is_mandatory:
            rules.append({
                "dimension": "Completeness",
                "rule_statement": f"{column_name} must not be blank",
            })
        return rules

    if constraint_lower == "number":
        rules.append({"dimension": "Validation", "rule_statement": f"{column_name} must be numeric"})
        if is_mandatory:
            rules.append({"dimension": "Completeness", "rule_statement": f"{column_name} must not be blank"})
        return rules

    if "number without thousand separator" in constraint_lower:
        rules.append({"dimension": "Validation", "rule_statement": f"{column_name} must be numeric"})
        rules.append({
            "dimension": "Conformity",
            "rule_statement": f"{column_name} must not contain thousand separators (commas)",
        })
        if is_mandatory:
            rules.append({"dimension": "Completeness", "rule_statement": f"{column_name} must not be blank"})
        return rules

    if "text" in constraint_lower or "string" in constraint_lower:
        rules.append({"dimension": "Validation", "rule_statement": f"{column_name} must contain valid text"})
        if is_mandatory:
            rules.append({"dimension": "Completeness", "rule_statement": f"{column_name} must not be blank"})
        return rules

    if not rules:
        rules.append({
            "dimension": "Conformity",
            "rule_statement": f"{column_name} must conform to format: {constraint}",
        })
        if is_mandatory:
            rules.append({"dimension": "Completeness", "rule_statement": f"{column_name} must not be blank"})
    return rules


def _find_matching_column(header_name: str, df: pd.DataFrame) -> Optional[str]:
    if df is None:
        return None
    header_lower = header_name.lower().strip()
    for col in df.columns:
        if str(col).lower().strip() == header_lower:
            return str(col)
    for col in df.columns:
        c = str(col).lower()
        if header_lower in c or c in header_lower:
            return str(col)
    return None


def extract_client_rules_from_excel_metadata(
    file_path: Optional[str], sheet_name: Optional[str], df: Optional[pd.DataFrame],
) -> List[Dict]:
    """Read raw Excel rows 1-4 to extract client-defined rules.

    Row 2 = constraints, Row 3 = headers (with optional * for mandatory).
    """
    try:
        if not file_path or not os.path.exists(file_path):
            return []
        if not sheet_name:
            return []

        df_meta = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=4)
        if len(df_meta) < 4:
            return []

        constraints_row = df_meta.iloc[2]
        headers_row = df_meta.iloc[3]

        extracted: List[Dict] = []
        for col_idx in range(len(constraints_row)):
            constraint = constraints_row.iloc[col_idx]
            header = headers_row.iloc[col_idx]
            if pd.isna(constraint) or pd.isna(header):
                continue
            constraint_str = str(constraint).strip()
            header_str = str(header).strip()
            if (not constraint_str or not header_str
                    or constraint_str.lower() == "nan"
                    or header_str.lower() == "nan"):
                continue
            clean_column_name = header_str.lstrip("*").strip()
            is_mandatory = header_str.startswith("*")
            rules = _parse_constraint_to_rules(constraint_str, clean_column_name, is_mandatory)
            for rule in rules:
                actual_col = _find_matching_column(clean_column_name, df)
                extracted.append({
                    "S.No": len(extracted) + 1,
                    "Column": actual_col if actual_col else clean_column_name,
                    "Business Field": clean_column_name,
                    "Dimension": rule["dimension"],
                    "Data Quality Rule": rule["rule_statement"],
                    "Issues Found": 0,
                    "Issues Found Example": "Client provided rule - From Excel metadata",
                    "Source": "Client Extracted",
                    "Metadata_Row": 2,
                })
        return extracted
    except Exception as exc:
        logger.warning("Could not extract client rules: %s", exc)
        return []
