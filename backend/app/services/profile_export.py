"""1:1 port of the Excel/JSON report exporters from features/profiling/ui.py.

Mirrors:
- _analyze_special_chars_detailed (lines 2518-2532)
- _apply_dq_rules_styling          (lines 2535-2622)
- _generate_excel_report           (lines 2625-3050)  → build_excel_report()
- _generate_json_report            (lines 3053-3105)  → build_json_report()
"""
from __future__ import annotations

import io
import json
import os
import re
import unicodedata
from collections import Counter
from datetime import datetime
from itertools import combinations
from typing import Any, Dict, List, Optional

import pandas as pd

from .ai_validation_engine import AIValidationEngine
from .match_rules import generate_match_rules, get_duplicate_count_values


def _analyze_special_chars_detailed(df: pd.DataFrame) -> List[Dict[str, Any]]:
    data: List[Dict[str, Any]] = []
    for col in df.select_dtypes(include=["object"]).columns:
        counter: Counter = Counter()
        for val in df[col].dropna().astype(str):
            for char in set(val):
                if ord(char) > 127 or (not char.isalnum() and not char.isspace()):
                    counter[char] += val.count(char)
        for char, count in counter.most_common():
            try:
                uname = unicodedata.name(char)
            except ValueError:
                uname = "UNKNOWN"
            data.append({"Column": col, "Character": char, "Unicode Name": uname, "Count": count})
    return data


def _apply_dq_rules_styling(workbook, worksheet, validation_df: pd.DataFrame) -> None:
    """Verbatim port of features/profiling/ui.py:2535-2622."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    dimension_colors = {
        "Accuracy": "DBEAFE", "Completeness": "DCFCE7", "Standardisation": "FEF3C7",
        "Validation": "FCE7F3", "Uniqueness": "F3E8FF", "Timeliness": "CCFBF1",
        "Integrity": "FEE2E2", "Conformity": "E0E7FF", "Reliability": "FFEDD5",
        "Relevance": "ECFCCB", "Precision": "FAE8FF", "Accessibility": "E0F2FE",
        "Character Length": "FDE68A",
        # Pre-rename names — keep export pretty for old profiles.
        "Consistency": "FEF3C7",
        "Validity": "FCE7F3",
    }
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, _ in enumerate(validation_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, row in enumerate(validation_df.itertuples(index=False), 2):
        dimension = row[2] if len(row) > 2 else ""
        issues_found = row[4] if len(row) > 4 else 0
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num == 3:
                color = dimension_colors.get(str(dimension), "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True)
            if col_num == 5:
                try:
                    issue_count = int(issues_found) if isinstance(issues_found, (int, float)) else 0
                    if issue_count > 0:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(color="991B1B", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(color="166534", bold=True)
                except Exception:
                    pass
            if col_num == 6 and isinstance(value, str):
                if value.startswith("All values valid") or value.startswith("Client provided"):
                    cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    cell.font = Font(color="166534", italic=True, size=10)
                elif value:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    cell.font = Font(color="92400E", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    column_widths = {"A": 6, "B": 20, "C": 18, "D": 35, "E": 12, "F": 40}
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    worksheet.row_dimensions[1].height = 25
    worksheet.freeze_panes = "A2"


def _is_valid_value(val: Any) -> bool:
    if pd.isna(val):
        return False
    str_val = str(val).strip().upper()
    return str_val not in ["", "NULL", "NONE", "NA", "N/A", "NAN", "BLANK", "-", "--", "---"]


def _has_valid_values(df: pd.DataFrame, indices: List, columns: List[str]) -> List:
    return [idx for idx in indices if all(_is_valid_value(df.loc[idx, c]) for c in columns)]


def _categorize_columns(df: pd.DataFrame) -> Dict[str, List[str]]:
    categories: Dict[str, List[str]] = {
        "unique_identifiers": [], "critical_business": [],
        "descriptive": [], "metadata": [],
    }
    for col in df.columns:
        col_lower = str(col).lower()
        series = df[col]
        unique_pct = (series.nunique() / len(series)) * 100 if len(series) > 0 else 0
        if any(kw in col_lower for kw in ["_id", "id", "key", "uuid", "guid", "serial", "sequence"]) and unique_pct > 95:
            categories["unique_identifiers"].append(col)
        elif any(kw in col_lower for kw in ["name", "number", "email", "phone", "mobile", "tax",
                                             "pan", "gstin", "account", "code", "vendor",
                                             "supplier", "customer"]):
            categories["critical_business"].append(col)
        elif any(kw in col_lower for kw in ["batch", "import", "action", "status", "created",
                                             "updated", "modified", "date", "time", "user"]):
            categories["metadata"].append(col)
        else:
            categories["descriptive"].append(col)
    return categories


def build_excel_report(
    df: pd.DataFrame,
    profiles: Dict[str, Any],
    filename: str,
    unified_validation_rules: Optional[pd.DataFrame] = None,
) -> tuple[bytes, str]:
    """Returns (bytes, filename) of multi-sheet Excel report."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) Executive Summary
        total_missing = sum(p.null_count for p in profiles.values())
        total_cells = len(df) * len(df.columns)
        quality = (sum(getattr(p, "non_null_percentage", 100) for p in profiles.values())
                   / len(profiles)) if profiles else 0
        completeness = ((total_cells - total_missing) / total_cells * 100) if total_cells else 0
        pd.DataFrame({
            "Metric": ["Total Rows", "Total Columns", "Total Cells", "Missing Cells",
                       "Completeness %", "Quality Score", "Generated At"],
            "Value": [
                len(df), len(df.columns), total_cells, total_missing,
                f"{completeness:.2f}%" if total_cells else "N/A",
                f"{quality:.1f}%",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        }).to_excel(writer, sheet_name="Executive Summary", index=False)

        # 2) Column Profiles
        profile_data: List[Dict[str, Any]] = []
        for col, p in profiles.items():
            dup_count = p.total_rows - p.unique_count
            profile_data.append({
                "Column Name": col,
                "Data Type": p.dtype,
                "Total Rows": p.total_rows,
                "Non-Null Count": p.total_rows - p.null_count,
                "Null Count": p.null_count,
                "Null Percentage": f"{p.null_percentage:.2f}%",
                "Unique Count": p.unique_count,
                "Duplicate Count": dup_count,
                "Duplicate Count Values": get_duplicate_count_values(df, col, max_items=None),
                "Unique Percentage": f"{p.unique_percentage:.2f}%",
                "Min Length": getattr(p, "min_length", "N/A"),
                "Max Length": getattr(p, "max_length", "N/A"),
                "Avg Length": f"{getattr(p, 'avg_length', 0):.2f}",
                "Risk Level": getattr(p, "risk_level", "Low"),
                "Risk Score": getattr(p, "risk_score", 0),
            })
        pd.DataFrame(profile_data).to_excel(writer, sheet_name="Column Profiles", index=False)

        # 3) Special Characters
        chars = _analyze_special_chars_detailed(df)
        (pd.DataFrame(chars) if chars else
         pd.DataFrame({"Message": ["No special characters found"]})).to_excel(
            writer, sheet_name="Special Characters", index=False)

        # 4) Match Rules
        match_rules = generate_match_rules(df, profiles)
        pd.DataFrame(match_rules).to_excel(writer, sheet_name="Match Rules", index=False)

        # 5) DQ Summary (if unified rules generated)
        if unified_validation_rules is not None and not unified_validation_rules.empty:
            dq_summary = unified_validation_rules.groupby("Dimension").agg({
                "S.No": "count", "Issues Found": "sum",
            }).reset_index()
            dq_summary.columns = ["Dimension", "Rule Count", "Total Issues"]
            dq_summary = dq_summary.sort_values("Rule Count", ascending=False)
            dq_summary.to_excel(writer, sheet_name="DQ Summary", index=False)

        # 6) Duplicates (exact + fuzzy similar)
        from utils.fuzzy_matching import FuzzyMatcher
        all_duplicate_records: List[Dict[str, Any]] = []
        fuzzy_threshold = 85.0

        for col in df.columns:
            try:
                if not pd.api.types.is_string_dtype(df[col]) and df[col].dtype != "object":
                    continue
                unique_values = df[col].dropna().unique()
                if len(unique_values) < 2:
                    continue
                value_counts = df[col].value_counts(dropna=True)
                duplicate_values = value_counts[value_counts > 1]
                processed_values: set = set()

                for dup_val in duplicate_values.index:
                    if isinstance(dup_val, str):
                        dup_str = str(dup_val).strip().upper()
                        if dup_str in ["", "NULL", "NONE", "NA", "N/A", "NAN", "BLANK",
                                       "-", "--", "---"]:
                            continue
                    matching_rows = df[df[col] == dup_val]
                    for _, row in matching_rows.iterrows():
                        record = {orig_col: "Missing" for orig_col in df.columns}
                        record[col] = dup_val
                        all_duplicate_records.append(record)
                    processed_values.add(str(dup_val))

                if len(unique_values) <= 1000:
                    matcher = FuzzyMatcher(algorithm="rapidfuzz", threshold=fuzzy_threshold)
                    for i, val1 in enumerate(unique_values):
                        if pd.isna(val1):
                            continue
                        val1_str = str(val1).strip()
                        if val1_str in processed_values:
                            continue
                        val1_upper = val1_str.upper()
                        if val1_upper in ["", "NULL", "NONE", "NA", "N/A", "NAN", "BLANK",
                                          "-", "--", "---"]:
                            continue
                        similar_group = [val1]
                        for j in range(i + 1, len(unique_values)):
                            val2 = unique_values[j]
                            if pd.isna(val2):
                                continue
                            val2_str = str(val2).strip()
                            if val2_str in processed_values:
                                continue
                            val2_upper = val2_str.upper()
                            if val2_upper in ["", "NULL", "NONE", "NA", "N/A", "NAN", "BLANK",
                                              "-", "--", "---"]:
                                continue
                            try:
                                similarity = matcher._rapidfuzz_match(
                                    matcher._normalize_text(val1_str),
                                    matcher._normalize_text(val2_str),
                                )
                                if similarity >= fuzzy_threshold:
                                    similar_group.append(val2)
                                    processed_values.add(val2_str)
                            except Exception:
                                continue
                        if len(similar_group) > 1:
                            processed_values.add(val1_str)
                            for similar_val in similar_group:
                                matching_rows = df[df[col] == similar_val]
                                for _, row in matching_rows.iterrows():
                                    record = {orig_col: "Missing" for orig_col in df.columns}
                                    record[col] = similar_val
                                    all_duplicate_records.append(record)
            except Exception:
                continue

        if all_duplicate_records:
            duplicates_df = pd.DataFrame(all_duplicate_records)
            duplicates_df = duplicates_df.sort_values(by=df.columns.tolist())
            duplicates_df.to_excel(writer, sheet_name="Duplicates", index=False)
        else:
            pd.DataFrame({
                "Message": ["No duplicate or similar values found (nulls/empty excluded)"],
            }).to_excel(writer, sheet_name="Duplicates", index=False)

        # 7) Exact Duplicate (5-strategy matrix)
        col_categories = _categorize_columns(df)
        all_duplicate_results: List[tuple] = []
        processed_indices: set = set()
        all_comparison_cols = [c for c in df.columns if c not in col_categories["unique_identifiers"]]
        if not all_comparison_cols:
            all_comparison_cols = df.columns.tolist()
        important_cols = col_categories["critical_business"] + col_categories["descriptive"][:5]

        # S1: single
        for col in important_cols:
            col_dups = df[df.duplicated(subset=[col], keep=False)]
            if len(col_dups) > 0:
                valid = _has_valid_values(df, col_dups.index.tolist(), [col])
                new_idx = [i for i in valid if i not in processed_indices]
                if new_idx:
                    new_dups = df.loc[new_idx].copy().sort_values(by=col)
                    new_dups.insert(0, "Matched_Fields", col)
                    all_duplicate_results.append((f"Single: {col}", new_dups))
                    processed_indices.update(new_idx)

        # S2: pairs
        if len(important_cols) >= 2:
            for c1, c2 in combinations(important_cols[:8], 2):
                two_dups = df[df.duplicated(subset=[c1, c2], keep=False)]
                if len(two_dups) > 0:
                    valid = _has_valid_values(df, two_dups.index.tolist(), [c1, c2])
                    new_idx = [i for i in valid if i not in processed_indices]
                    if new_idx:
                        new_dups = df.loc[new_idx].copy().sort_values(by=[c1, c2])
                        new_dups.insert(0, "Matched_Fields", f"{c1}, {c2}")
                        all_duplicate_results.append((f"Pair: {c1}+{c2}", new_dups))
                        processed_indices.update(new_idx)

        # S3: triples
        if len(important_cols) >= 3:
            for c1, c2, c3 in combinations(important_cols[:6], 3):
                three_dups = df[df.duplicated(subset=[c1, c2, c3], keep=False)]
                if len(three_dups) > 0:
                    valid = _has_valid_values(df, three_dups.index.tolist(), [c1, c2, c3])
                    new_idx = [i for i in valid if i not in processed_indices]
                    if new_idx:
                        new_dups = df.loc[new_idx].copy().sort_values(by=[c1, c2, c3])
                        new_dups.insert(0, "Matched_Fields", f"{c1}, {c2}, {c3}")
                        all_duplicate_results.append((f"Triple: {c1}+{c2}+{c3}", new_dups))
                        processed_indices.update(new_idx)

        # S4: critical business
        if col_categories["critical_business"]:
            biz_dups = df[df.duplicated(subset=col_categories["critical_business"], keep=False)]
            if len(biz_dups) > 0:
                valid = _has_valid_values(df, biz_dups.index.tolist(),
                                          col_categories["critical_business"])
                new_idx = [i for i in valid if i not in processed_indices]
                if new_idx:
                    new_dups = df.loc[new_idx].copy().sort_values(
                        by=col_categories["critical_business"][:3])
                    new_dups.insert(0, "Matched_Fields", ", ".join(col_categories["critical_business"]))
                    all_duplicate_results.append(("Business Keys", new_dups))
                    processed_indices.update(new_idx)

        # S5: complete row
        complete_dups = df[df.duplicated(subset=all_comparison_cols, keep=False)]
        if len(complete_dups) > 0:
            valid = _has_valid_values(df, complete_dups.index.tolist(), all_comparison_cols)
            new_idx = [i for i in valid if i not in processed_indices]
            if new_idx:
                new_dups = df.loc[new_idx].copy().sort_values(by=all_comparison_cols[:3])
                new_dups.insert(0, "Matched_Fields", "All fields")
                all_duplicate_results.append(("Complete Match", new_dups))
                processed_indices.update(new_idx)

        if all_duplicate_results:
            combined = [t[1] for t in all_duplicate_results]
            final = pd.concat(combined, ignore_index=True)
            final.to_excel(writer, sheet_name="Exact Duplicate", index=False)
        else:
            pd.DataFrame({
                "Message": ["No duplicates detected (null/empty values excluded)"],
                "Detection_Method": ["Partial & Complete match analysis"],
                "Columns_Analyzed": [f"{len(all_comparison_cols)} columns"],
            }).to_excel(writer, sheet_name="Exact Duplicate", index=False)

    output.seek(0)
    base_name = os.path.splitext(str(filename or "dataset"))[0]
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", base_name)
    fname = f"{safe}_Data_Profile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return output.getvalue(), fname


def build_json_report(df: pd.DataFrame, profiles: Dict[str, Any], filename: str) -> tuple[bytes, str]:
    """Verbatim port of features/profiling/ui.py:3053-3105."""
    report = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "ai_powered": True,
            "dq_dimensions": AIValidationEngine.DQ_DIMENSIONS,
        },
        "executive_summary": {
            "missing_cells": sum(p.null_count for p in profiles.values()),
            "completeness": (f"{sum(getattr(p, 'non_null_percentage', 100) for p in profiles.values()) / len(profiles):.2f}%"
                             if profiles else "N/A"),
        },
        "column_profiles": {
            col: {
                "type": p.dtype,
                "null_percentage": p.null_percentage,
                "unique_percentage": p.unique_percentage,
                "duplicate_count": p.total_rows - p.unique_count,
            } for col, p in profiles.items()
        },
        "match_rules": generate_match_rules(df, profiles),
    }
    json_data = json.dumps(report, indent=2, default=str)
    base_name = os.path.splitext(str(filename or "dataset"))[0]
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", base_name)
    fname = f"{safe}_Data_Profile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return json_data.encode("utf-8"), fname
